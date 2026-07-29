import json
import re
from html import unescape
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from urllib.error import URLError, HTTPError
from urllib.request import Request, urlopen

from openpyxl import load_workbook


BASE = Path(__file__).resolve().parent
OUT = BASE / "order_report_from_raw_data.html"
PRODUCT_FILE = Path("/Users/nguyenhungvi/Downloads/Product Haravan.xlsx")
PRIORITY_FILE = Path("/Users/nguyenhungvi/Downloads/SKU Priority.xlsx")
IMAGE_LINK_FILE = Path("/Users/nguyenhungvi/Downloads/link hình ảnh.xlsx")
IMAGE_LINK_UPDATE_FILE = Path("/Users/nguyenhungvi/Downloads/ link ảnh update.xlsx")
IMAGE_CACHE = BASE / "product_image_cache.json"
LOCAL_IMAGE_DIR = BASE / "product_images"
FETCH_PRODUCT_IMAGES = False
DOWNLOAD_IMAGE_FILES = True


def as_number(value):
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    try:
        return float(text)
    except ValueError:
        return 0.0


def as_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date().isoformat()
    except ValueError:
        return None


def month_key(filename):
    match = re.search(r"Orders_T(\d+)_(\d{4})\.xlsx$", filename.name)
    if not match:
        return (9999, 99)
    return (int(match.group(2)), int(match.group(1)))


def clean_text(value, fallback="Chưa phân loại"):
    if value in (None, ""):
        return fallback
    return str(value).strip() or fallback


def infer_group(product):
    name = clean_text(product)
    if "Thank You Card" in name:
        return "POSM / Thank You Card"
    for brand in (" BlueStone", " Carez", " Bluestone"):
        if brand in name:
            return name.split(brand)[0].strip() or "Khác"
    name = re.sub(r"\s+[A-Z]{2,}[-A-Z0-9]*\d[\w-]*.*$", "", name).strip()
    name = re.sub(r"\s+\d+(\.\d+)?\s*(Lít|W|ml|Kg|lít|w|ML).*$", "", name).strip()
    return name or "Khác"


def extract_model(product_name):
    text = clean_text(product_name, "")
    patterns = [
        r"\b([A-Z]{2,5}-\d{3,5}[A-Z]?)\b",
        r"\b([A-Z]{2,5}\d{3,5}[A-Z]?)\b",
        r"\b([A-Z]{2,5}-[A-Z]{2,5}-\d{3,5}[A-Z]?)\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return re.sub(r"[^A-Z0-9]", "", match.group(1).upper())
    return ""


def normalize_url(url):
    if url in (None, ""):
        return ""
    text = str(url).strip()
    if not text or text.startswith("#") or text.upper() in {"#REF!", "#N/A", "#VALUE!", "#NAME?"}:
        return ""
    if text.startswith("//"):
        return "https:" + text
    if text.startswith("http://"):
        return "https://" + text[len("http://") :]
    return text


def local_image_name(sku, url):
    suffix = Path(url.split("?", 1)[0]).suffix.lower()
    if suffix not in {".jpg", ".jpeg", ".png", ".webp", ".gif"}:
        suffix = ".jpg"
    safe_sku = re.sub(r"[^A-Za-z0-9_-]+", "_", sku)
    return f"{safe_sku}{suffix}"


def download_image_file(sku, url):
    if not DOWNLOAD_IMAGE_FILES or not sku or not url:
        return ""
    LOCAL_IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    filename = local_image_name(sku, url)
    destination = LOCAL_IMAGE_DIR / filename
    if destination.exists() and destination.stat().st_size > 0:
        return f"product_images/{filename}"
    try:
        req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urlopen(req, timeout=30) as response:
            data = response.read()
        if data:
            destination.write_bytes(data)
            return f"product_images/{filename}"
    except (URLError, HTTPError, TimeoutError, OSError):
        return ""
    return ""


def extract_image_url(page_html):
    patterns = [
        r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\']([^"\']+)',
        r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+property=["\']og:image["\']',
        r'"image"\s*:\s*"([^"]+)"',
    ]
    for pattern in patterns:
        match = re.search(pattern, page_html, re.IGNORECASE)
        if match:
            return normalize_url(unescape(match.group(1)))
    return ""


def fetch_product_image(url):
    if not url:
        return ""
    try:
        req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urlopen(req, timeout=20) as response:
            page_html = response.read(500_000).decode("utf-8", "ignore")
        return extract_image_url(page_html)
    except (URLError, HTTPError, TimeoutError, OSError):
        return ""


def load_product_map():
    product_map = {}
    if not PRODUCT_FILE.exists():
        return product_map

    image_cache = {}
    if IMAGE_CACHE.exists():
        try:
            image_cache = json.loads(IMAGE_CACHE.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            image_cache = {}

    wb = load_workbook(PRODUCT_FILE, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {str(header).strip(): pos for pos, header in enumerate(headers) if header is not None}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = clean_text(row[idx["Barcode"]], "")
        if not sku:
            continue
        url = normalize_url(row[idx["Url"]])
        image_url = image_cache.get(url, "")
        if FETCH_PRODUCT_IMAGES and url and not image_url:
            image_url = fetch_product_image(url)
            image_cache[url] = image_url
        product_map[sku] = {
            "url": url,
            "image": image_url,
            "product": clean_text(row[idx.get("Tên Sản Phẩm", idx["Barcode"])], ""),
        }

    IMAGE_CACHE.write_text(json.dumps(image_cache, ensure_ascii=False, indent=2), encoding="utf-8")
    return product_map


def load_image_link_workbook(path):
    image_map = {}
    if not path.exists():
        return image_map

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {header: pos for pos, header in enumerate(headers) if header}
    sku_col = idx.get("barcode")
    if sku_col is None:
        sku_col = idx.get("sku")
    image_col = idx.get("link image")
    if image_col is None:
        image_col = idx.get("link ảnh")
    if image_col is None:
        image_col = idx.get("image")
    if image_col is None:
        image_col = idx.get("link")
    if sku_col is None or image_col is None:
        return image_map

    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = clean_text(row[sku_col] if sku_col < len(row) else None, "")
        image_url = normalize_url(row[image_col] if image_col < len(row) else None)
        if sku and image_url:
            image_map[sku] = download_image_file(sku, image_url) or image_url
    return image_map


def load_image_link_map():
    image_map = {}
    for path in [IMAGE_LINK_FILE, IMAGE_LINK_UPDATE_FILE]:
        image_map.update(load_image_link_workbook(path))
    return image_map


def build_model_index(mapping):
    model_index = {}
    for sku, info in mapping.items():
        candidates = [sku, info.get("product", "")]
        for candidate in candidates:
            model = extract_model(candidate)
            if model:
                model_index.setdefault(model, sku)
    return model_index


def find_header_row(ws, required):
    required = set(required)
    for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
        labels = {str(value).strip() for value in row if value is not None}
        if required.issubset(labels):
            return row_number, [str(value).strip() if value is not None else "" for value in row]
    return None, []


def load_priority_map():
    priority_map = {}
    if not PRIORITY_FILE.exists():
        return priority_map

    wb = load_workbook(PRIORITY_FILE, read_only=True, data_only=True)
    for ws in wb.worksheets:
        header_row, headers = find_header_row(ws, ["SKU"])
        if not header_row:
            continue
        idx = {header: pos for pos, header in enumerate(headers) if header}
        sku_col = idx.get("SKU")
        priority_col = idx.get("Classify")
        pm_classify_col = idx.get("PM Classify")
        group_col = idx.get("Group")
        product_col = idx.get("Product Name")
        image_col = idx.get("Image")
        if sku_col is None or priority_col is None:
            continue

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            sku = clean_text(row[sku_col] if sku_col < len(row) else None, "")
            priority = clean_text(row[priority_col] if priority_col < len(row) else None, "")
            if not sku or not priority.startswith("Priority"):
                continue
            priority_map[sku] = {
                "priority": priority,
                "classify": clean_text(row[pm_classify_col] if pm_classify_col is not None and pm_classify_col < len(row) else None, "Chưa phân loại"),
                "group": clean_text(row[group_col] if group_col is not None and group_col < len(row) else None, ""),
                "product": clean_text(row[product_col] if product_col is not None and product_col < len(row) else None, ""),
                "image": normalize_url(row[image_col]) if image_col is not None and image_col < len(row) else "",
            }
    return priority_map


def read_records():
    raw = []
    order_buckets = defaultdict(list)
    product_map = load_product_map()
    priority_map = load_priority_map()
    image_link_map = load_image_link_map()
    product_model_index = build_model_index(product_map)
    priority_model_index = build_model_index(priority_map)
    files = sorted(BASE.glob("Orders_T*.xlsx"), key=month_key)

    for file in files:
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {header: pos for pos, header in enumerate(headers)}

        for row in ws.iter_rows(min_row=2, values_only=True):
            order_id = clean_text(row[idx["Mã đơn hàng"]], "")
            date = as_date(row[idx["Ngày đặt hàng"]])
            if not order_id or not date:
                continue

            qty = as_number(row[idx["Số lượng sản phẩm"]])
            price = as_number(row[idx["Giá sản phẩm"]])
            gross_revenue = qty * price
            barcode = clean_text(row[idx["Mã sản phẩm"]], "Không có barcode")
            product = clean_text(row[idx["Tên sản phẩm"]], "Không có tên sản phẩm")
            model = extract_model(product)
            sku = (
                priority_model_index.get(model)
                or product_model_index.get(model)
                or (barcode if barcode in priority_map or barcode in product_map else barcode)
            )
            product_info = product_map.get(sku, {})
            priority_info = priority_map.get(sku, {})
            product = priority_info.get("product") or product_info.get("product") or product
            group = priority_info.get("group") or infer_group(product)
            channel = clean_text(row[idx["Kênh bán hàng"]], "Không xác định").lower()
            cancel = clean_text(row[idx["Trạng thái hủy"]], "No")
            payment = clean_text(row[idx["Phương thức thanh toán"]], "Không xác định")
            priority = priority_info.get("priority", "Others")
            classify = priority_info.get("classify", "Chưa phân loại")
            image_url = image_link_map.get(sku) or priority_info.get("image") or product_info.get("image", "")
            product_url = product_info.get("url", "")

            item = {
                "d": date,
                "m": date[:7],
                "c": channel,
                "x": cancel,
                "q": round(qty, 4),
                "r": 0,
                "_gross": gross_revenue,
                "_order_total": as_number(row[idx["Tổng cộng"]]),
                "s": sku,
                "v": barcode,
                "p": product,
                "g": group,
                "i": image_url,
                "u": product_url,
                "ks": priority,
                "cl": classify,
                "o": order_id,
                "pm": payment,
                "ship": round(as_number(row[idx["Phí vận chuyển"]]), 2),
                "pay": clean_text(row[idx["Tình trạng thanh toán"]], "Không xác định"),
                "ful": clean_text(row[idx["Tình trạng giao hàng"]], "Không xác định"),
            }
            raw.append(item)
            order_buckets[order_id].append(item)

    for order_items in order_buckets.values():
        order_total = order_items[0].get("_order_total", 0)
        gross_total = sum(item.get("_gross", 0) for item in order_items)
        if gross_total:
            allocated = 0
            for item in order_items[:-1]:
                item["r"] = round(order_total * item.get("_gross", 0) / gross_total, 2)
                allocated += item["r"]
            order_items[-1]["r"] = round(order_total - allocated, 2)
        elif order_items:
            even_share = round(order_total / len(order_items), 2)
            allocated = even_share * (len(order_items) - 1)
            for item in order_items[:-1]:
                item["r"] = even_share
            order_items[-1]["r"] = round(order_total - allocated, 2)

    for item in raw:
        item.pop("_gross", None)
        item.pop("_order_total", None)
    return raw


def build_report_data(records):
    dates = sorted({item["d"] for item in records})
    channels = sorted({item["c"] for item in records})
    groups = sorted({item["g"] for item in records})
    skus = sorted({item["s"] for item in records})
    priorities = ["Priority 1", "Priority 2", "Priority 3", "Priority 4", "Others"]
    classifies = sorted({item["cl"] for item in records})
    default_to = dates[-1]
    default_from = default_to[:8] + "01"
    return {
        "records": records,
        "meta": {
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "recordCount": len(records),
            "minDate": dates[0],
            "maxDate": dates[-1],
            "defaultFrom": default_from,
            "defaultTo": default_to,
            "channels": channels,
            "groups": groups,
            "skus": skus,
            "keySummers": priorities,
            "classifies": classifies,
            "imageAssets": {},
        },
    }

def html_template(report_json):
    return """<!doctype html>
<html lang="vi">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Order Haravan Report - Raw Data</title>
  <style>
    :root {
      --bg: #f3f7fb; --card: #ffffff; --ink: #16324f; --muted: #60758d; --line: #dbe6f0;
      --blue: #4d89e8; --orange: #f39a3f; --green: #1a9b5f; --red: #db4d4d;
      --shadow: 0 10px 30px rgba(15, 55, 95, 0.08); --radius: 18px;
    }
    * { box-sizing: border-box; }
    body { margin: 0; font-family: "Segoe UI", "Helvetica Neue", Arial, sans-serif; color: var(--ink); background: #fff; }
    .page { max-width: 1500px; margin: 0 auto; padding: 18px 18px 60px; }
    .hero, .filter-card, .panel, .metric { background: var(--card); border: 1px solid rgba(23,71,115,.08); border-radius: var(--radius); box-shadow: var(--shadow); }
    .hero { border-radius: 26px; padding: 22px; margin-bottom: 18px; border-color: rgba(77,137,232,.15); }
    .hero h1 { margin: 0; font-size: 30px; line-height: 1.1; }
    .hero p { margin: 8px 0 0; color: var(--muted); font-size: 14px; }
    .filters { display: grid; grid-template-columns: repeat(5, minmax(0, 1fr)); gap: 12px; margin-bottom: 18px; }
    .filter-card { padding: 12px 14px; min-height: 80px; }
    label, .toolbar-label { display: block; font-size: 11px; font-weight: 800; text-transform: uppercase; letter-spacing: .03em; color: var(--muted); margin-bottom: 7px; }
    input, select, .multi-select-trigger, .download-btn { width: 100%; border: 1px solid var(--line); border-radius: 12px; padding: 10px 12px; font-size: 14px; color: var(--ink); background: #f9fbfd; }
    .download-btn { width: auto; font-weight: 800; cursor: pointer; white-space: nowrap; }
    .filter-hint, .subtle, .note, .panel-subtitle { color: var(--muted); font-size: 12px; }
    .multi-select { position: relative; }
    .multi-select-trigger { min-height: 44px; text-align: left; cursor: pointer; }
    .multi-select-menu { position: absolute; top: calc(100% + 8px); left: 0; right: 0; z-index: 20; display: none; background: #fff; border: 1px solid #dbe6f0; border-radius: 16px; box-shadow: 0 18px 40px rgba(15,55,95,.16); padding: 12px; }
    .multi-select.open .multi-select-menu { display: block; }
    .multi-select-search { margin-bottom: 8px; }
    .multi-select-actions { display: flex; justify-content: space-between; gap: 8px; margin-bottom: 8px; }
    .multi-select-actions button { border: none; background: transparent; color: var(--blue); font-size: 12px; font-weight: 800; cursor: pointer; padding: 0; }
    .multi-select-options { max-height: 220px; overflow: auto; display: grid; gap: 6px; }
    .multi-select-option { display: flex; align-items: center; gap: 10px; padding: 8px 10px; border-radius: 10px; cursor: pointer; font-size: 13px; }
    .multi-select-option:hover { background: #f4f8fd; }
    .multi-select-option input { width: auto; margin: 0; }
    .selected-chips { display: flex; flex-wrap: wrap; gap: 6px; margin-top: 8px; min-height: 24px; }
    .selected-chip { display: inline-flex; align-items: center; gap: 6px; padding: 4px 8px; border-radius: 999px; background: #eef5fd; font-size: 11px; font-weight: 700; }
    .selected-chip button { border: none; background: transparent; color: var(--muted); cursor: pointer; padding: 0; }
    .section-title { margin: 18px 0 12px; font-size: 18px; font-weight: 800; text-decoration: underline; text-underline-offset: 4px; }
    .comparison-note { margin: -4px 0 12px; padding: 10px 14px; border: 1px solid #dbe7f4; border-radius: 12px; background: #f7faff; color: var(--muted); font-size: 13px; font-weight: 600; }
    .comparison-note strong { color: var(--ink); }
    .metrics { display: grid; grid-template-columns: repeat(4, minmax(0, 1fr)); gap: 14px; margin-bottom: 18px; }
    .metric { padding: 14px 16px; border-color: rgba(77,137,232,.25); }
    .metric .label { color: var(--muted); font-size: 14px; text-transform: uppercase; letter-spacing: .03em; }
    .metric .value { margin-top: 4px; font-size: 34px; font-weight: 800; line-height: 1; }
    .delta { margin-top: 8px; font-size: 14px; font-weight: 700; }
    .delta.up { color: var(--green); } .delta.down { color: var(--red); } .delta.flat { color: var(--muted); }
    .breakdown-grid { display: grid; grid-template-columns: minmax(250px,.8fr) minmax(0,1.7fr); gap: 16px; margin-bottom: 18px; }
    .grid-2 { display: grid; grid-template-columns: 2fr 1fr; gap: 16px; margin-bottom: 18px; }
    .panel { padding: 16px; overflow: hidden; }
    .panel h3 { margin: 0 0 14px; font-size: 20px; }
    .panel-head, .panel-toolbar, .raw-toolbar { display: flex; align-items: center; justify-content: space-between; gap: 12px; margin-bottom: 14px; flex-wrap: wrap; }
    .view-switch { display: inline-flex; flex-wrap: wrap; gap: 8px; margin-bottom: 14px; }
    .view-switch button { border: 1px solid #d7e4ef; background: #eef7f5; color: #527b75; border-radius: 8px; padding: 8px 14px; font-size: 13px; font-weight: 800; cursor: pointer; }
    .view-switch button.active { background: var(--blue); border-color: var(--blue); color: #fff; }
    .viz-list { display: grid; gap: 12px; }
    .viz-row { display: grid; grid-template-columns: minmax(120px,1.2fr) 3fr auto; gap: 12px; align-items: center; }
    .viz-name { font-size: 14px; font-weight: 700; line-height: 1.2; }
    .viz-bar { position: relative; height: 14px; border-radius: 999px; overflow: hidden; background: #edf3f9; }
    .viz-fill { position: absolute; inset: 0 auto 0 0; border-radius: inherit; }
    .viz-meta { text-align: right; font-size: 13px; font-weight: 700; color: var(--muted); white-space: nowrap; }
    .chart-shell { position: relative; width: 100%; overflow-x: auto; }
    svg { display: block; width: 100%; height: auto; }
    .legend { display: flex; gap: 14px; flex-wrap: wrap; margin-bottom: 10px; font-size: 14px; font-weight: 700; }
    .legend span { display: inline-flex; align-items: center; gap: 8px; }
    .legend i { display: inline-block; width: 14px; height: 14px; border-radius: 4px; }
    table { width: 100%; border-collapse: collapse; font-size: 13px; }
    th, td { padding: 10px; border-bottom: 1px solid #edf2f7; text-align: left; vertical-align: middle; }
    thead th { background: #eef5fd; font-size: 12px; text-transform: uppercase; letter-spacing: .03em; position: sticky; top: 0; z-index: 2; }
    tbody tr:nth-child(even) { background: #fafcff; }
    .table-scroll { max-height: 720px; overflow: auto; border: 1px solid #edf2f7; border-radius: 14px; background: #fff; }
    .sort-btn { display: inline-flex; align-items: center; gap: 6px; border: none; background: transparent; color: inherit; font: inherit; font-weight: 800; cursor: pointer; padding: 0; text-transform: inherit; letter-spacing: inherit; }
    .rank { color: var(--muted); width: 42px; }
    .sku-code { font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace; font-size: 13px; font-weight: 800; word-break: break-all; }
    .sku-thumb, .raw-thumb { width: 64px; height: 64px; object-fit: contain; border-radius: 12px; border: 1px solid #e5edf6; background: #fff; padding: 4px; }
    .raw-thumb { width: 72px; height: 72px; }
    .product-link { color: var(--blue); font-weight: 800; text-decoration: none; }
    .product-link:hover { text-decoration: underline; }
    .raw-table table { min-width: 1320px; }
    .raw-product { max-width: 280px; white-space: normal; line-height: 1.35; }
    .empty { padding: 30px 0; text-align: center; color: var(--muted); }
    .summary-total { margin-top: 12px; padding: 14px; border-radius: 16px; background: #fff; border: 1px solid #dde9f5; }
    .summary-total-grid { display: grid; grid-template-columns: 1.4fr repeat(3, minmax(0,1fr)); gap: 12px; align-items: center; }
    .summary-stat { background: #fff; border: 1px solid #edf3f9; border-radius: 12px; padding: 10px 12px; }
    .summary-stat-label { font-size: 11px; font-weight: 800; color: var(--muted); text-transform: uppercase; }
    .summary-stat-value { margin-top: 4px; font-size: 20px; font-weight: 800; line-height: 1.05; }
    @media (max-width: 1200px) { .filters, .metrics, .breakdown-grid, .grid-2 { grid-template-columns: 1fr 1fr; } }
    @media (max-width: 760px) { .filters, .metrics, .breakdown-grid, .grid-2, .summary-total-grid { grid-template-columns: 1fr; } .hero h1 { font-size: 24px; } .metric .value { font-size: 28px; } }
  </style>
</head>
<body>
  <div class="page">
    <section class="hero">
      <h1>Order Haravan Report</h1>
      <p id="heroMeta"></p>
    </section>
    <section class="filters">
      <div class="filter-card"><label for="fromDate">Từ ngày</label><input id="fromDate" type="date" /></div>
      <div class="filter-card"><label for="toDate">Đến ngày</label><input id="toDate" type="date" /></div>
      <div class="filter-card"><label>Kênh bán hàng</label><div class="multi-select" id="channelFilter"><button type="button" class="multi-select-trigger" id="channelTrigger"></button><div class="multi-select-menu"><input class="multi-select-search" id="channelSearch" placeholder="Tìm kênh..." /><div class="multi-select-actions"><button type="button" data-filter="channel" data-action="all">Chọn tất cả</button><button type="button" data-filter="channel" data-action="clear">Bỏ chọn</button></div><div class="multi-select-options" id="channelOptions"></div></div></div><div class="selected-chips" id="channelChips"></div><div class="filter-hint">Bấm chọn nhiều kênh cùng lúc</div></div>
      <div class="filter-card"><label>Group</label><div class="multi-select" id="groupFilter"><button type="button" class="multi-select-trigger" id="groupTrigger"></button><div class="multi-select-menu"><input class="multi-select-search" id="groupSearch" placeholder="Tìm group..." /><div class="multi-select-actions"><button type="button" data-filter="group" data-action="all">Chọn tất cả</button><button type="button" data-filter="group" data-action="clear">Bỏ chọn</button></div><div class="multi-select-options" id="groupOptions"></div></div></div><div class="selected-chips" id="groupChips"></div><div class="filter-hint">Group được suy ra từ tên sản phẩm</div></div>
      <div class="filter-card"><label>Barcode / SKU</label><div class="multi-select" id="skuFilter"><button type="button" class="multi-select-trigger" id="skuTrigger"></button><div class="multi-select-menu"><input class="multi-select-search" id="skuSearch" placeholder="Tìm SKU / barcode..." /><div class="multi-select-actions"><button type="button" data-filter="sku" data-action="all">Chọn tất cả</button><button type="button" data-filter="sku" data-action="clear">Bỏ chọn</button></div><div class="multi-select-options" id="skuOptions"></div></div></div><div class="selected-chips" id="skuChips"></div><div class="filter-hint">Tìm và tick nhiều SKU</div></div>
      <div class="filter-card"><label>Priority</label><div class="multi-select" id="keySummerFilter"><button type="button" class="multi-select-trigger" id="keySummerTrigger"></button><div class="multi-select-menu"><input class="multi-select-search" id="keySummerSearch" placeholder="Tìm priority..." /><div class="multi-select-actions"><button type="button" data-filter="keySummer" data-action="all">Chọn tất cả</button><button type="button" data-filter="keySummer" data-action="clear">Bỏ chọn</button></div><div class="multi-select-options" id="keySummerOptions"></div></div></div><div class="selected-chips" id="keySummerChips"></div></div>
      <div class="filter-card"><label>Classify</label><div class="multi-select" id="classifyFilter"><button type="button" class="multi-select-trigger" id="classifyTrigger"></button><div class="multi-select-menu"><input class="multi-select-search" id="classifySearch" placeholder="Tìm classify..." /><div class="multi-select-actions"><button type="button" data-filter="classify" data-action="all">Chọn tất cả</button><button type="button" data-filter="classify" data-action="clear">Bỏ chọn</button></div><div class="multi-select-options" id="classifyOptions"></div></div></div><div class="selected-chips" id="classifyChips"></div></div>
      <div class="filter-card"><label for="cancelFilter">Trạng thái hủy</label><select id="cancelFilter"><option value="all">Tất cả</option><option value="No">No</option><option value="Yes">Yes</option></select></div>
      <div class="filter-card"><label for="rawPeriodFilter">Raw date scope</label><select id="rawPeriodFilter"><option value="week">Tuần</option><option value="month">Tháng</option><option value="year">Năm</option><option value="day">Ngày</option></select></div>
    </section>

    <h2 class="section-title">1. Total DT / Volume</h2>
    <div class="comparison-note" id="comparisonNote"></div>
    <section class="metrics">
      <article class="metric"><div class="label">DT</div><div class="value" id="metricRevenue">-</div><div class="delta" id="metricRevenueDelta">-</div></article>
      <article class="metric"><div class="label">Volume</div><div class="value" id="metricVolume">-</div><div class="delta" id="metricVolumeDelta">-</div></article>
      <article class="metric"><div class="label">ASP</div><div class="value" id="metricAsp">-</div><div class="delta" id="metricAspDelta">-</div></article>
      <article class="metric"><div class="label">% Hủy</div><div class="value" id="metricCancel">-</div><div class="delta" id="metricCancelDelta">-</div></article>
    </section>

    <section class="panel">
      <div class="panel-head"><h3>DT và Volume theo ngày</h3><div class="view-switch" id="trendViewSwitch"><button type="button" data-trend-view="day" class="active">Ngày</button><button type="button" data-trend-view="week">Tuần</button><button type="button" data-trend-view="month">Tháng</button></div></div>
      <div class="legend"><span><i style="background:#4d89e8"></i> DT</span><span><i style="background:#f39a3f"></i> Volume</span></div>
      <div class="chart-shell" id="trendChart"></div>
    </section>

    <h2 class="section-title">2. Performance by Channel</h2>
    <section class="breakdown-grid"><article class="panel"><h3>Tỷ trọng doanh thu theo kênh</h3><div class="panel-subtitle">Top kênh đóng góp doanh thu trong kỳ đang chọn</div><div id="channelViz"></div></article><article class="panel"><h3>Tổng hợp kênh bán</h3><div id="channelTable"></div></article></section>
    <h2 class="section-title">3. Performance by Group</h2>
    <section class="breakdown-grid"><article class="panel"><h3>Tỷ trọng doanh thu theo group</h3><div class="panel-subtitle">Top group suy ra từ tên sản phẩm</div><div id="groupViz"></div></article><article class="panel"><div class="panel-head"><h3>Tổng hợp group</h3><button id="downloadGroupData" class="download-btn">Tải group data</button></div><div id="groupTable"></div></article></section>
    <h2 class="section-title">4. Performance by Priority / Classify</h2>
    <section class="breakdown-grid"><article class="panel"><h3>Priority</h3><div class="panel-subtitle">Priority lấy từ file SKU Priority; SKU ngoài danh sách được gán Others</div><div id="keySummerViz"></div></article><article class="panel"><h3>Tổng hợp Priority</h3><div id="keySummerTable"></div></article></section>
    <section class="breakdown-grid"><article class="panel"><h3>Classify</h3><div class="panel-subtitle">Classify suy theo DT, volume và ASP</div><div id="classifyViz"></div></article><article class="panel"><h3>Tổng hợp Classify</h3><div id="classifyTable"></div></article></section>

    <h2 class="section-title">5. Performance by SKU</h2>
    <section class="panel">
      <div class="panel-toolbar"><div class="subtle">Top SKU theo DT, có thể sắp xếp theo Barcode, sản phẩm, group, DT, volume, ASP.</div><button id="downloadSkuData" class="download-btn">Tải data</button></div>
      <h3>Top SKU theo DT</h3><div id="skuTable"></div>
      <section class="grid-2" style="margin-top:20px"><article class="panel"><h3>Top 20 Growth</h3><div class="panel-subtitle">SKU có DT quy đổi > 5 triệu / tuần</div><div id="skuGrowthTable"></div></article><article class="panel"><h3>Top 20 Reduce</h3><div class="panel-subtitle">SKU có DT quy đổi > 5 triệu / tuần</div><div id="skuReduceTable"></div></article></section>
      <div class="note">Nguồn dữ liệu: các file Orders_T*.xlsx trong folder hiện tại, Product Haravan.xlsx để join link/ảnh sản phẩm, và SKU Priority.xlsx để phân loại Priority. DT = Tổng cộng sau giảm giá theo mã đơn duy nhất; với đơn có nhiều sản phẩm, DT được phân bổ về từng dòng theo tỷ trọng Giá sản phẩm x Số lượng để tránh nhân đôi mã đơn. Dấu chấm/dấu phẩy số dùng locale vi-VN.</div>
    </section>

    <h2 class="section-title">6. Raw Data</h2>
    <section class="panel"><div class="raw-toolbar"><div class="subtle">Raw data được gom theo SKU + kênh + kỳ đang chọn.</div><button id="downloadRawData" class="download-btn">Tải raw data</button></div><h3>Raw data by period</h3><div id="rawTable"></div></section>
  </div>
  <script>
    const REPORT_DATA = __REPORT_JSON__;
    const PALETTE = ["#4d89e8","#f39a3f","#9d73db","#a9bf52","#2aa6b8","#de72aa","#dcb774","#59b0db","#f0c635","#7d8940","#a67a63","#9aa9c9"];
    const state = { from: REPORT_DATA.meta.defaultFrom, to: REPORT_DATA.meta.defaultTo, channel: [], group: [], sku: [], keySummer: [], classify: [], cancel: "all", rawPeriod: "week", trendView: "day", summarySort: { channel: {key:"revenue",dir:"desc"}, group: {key:"revenue",dir:"desc"}, keySummer: {key:"revenue",dir:"desc"}, classify: {key:"revenue",dir:"desc"} }, skuSort: {key:"revenue",dir:"desc"} };
    const multiSelectConfig = {
      channel: { values: REPORT_DATA.meta.channels, label: "kênh", allLabel: "Tất cả kênh" },
      group: { values: REPORT_DATA.meta.groups, label: "group", allLabel: "Tất cả group" },
      sku: { values: REPORT_DATA.meta.skus, label: "SKU", allLabel: "Tất cả SKU" },
      keySummer: { values: REPORT_DATA.meta.keySummers, label: "priority", allLabel: "Tất cả priority" },
      classify: { values: REPORT_DATA.meta.classifies, label: "classify", allLabel: "Tất cả classify" }
    };
    function setup() {
      document.getElementById("heroMeta").textContent = `Generated ${new Date(REPORT_DATA.meta.generatedAt).toLocaleString("vi-VN")} | ${REPORT_DATA.meta.recordCount.toLocaleString("vi-VN")} line items | ${REPORT_DATA.meta.minDate} -> ${REPORT_DATA.meta.maxDate}`;
      document.getElementById("fromDate").value = state.from; document.getElementById("toDate").value = state.to;
      ["channel","group","sku","keySummer","classify"].forEach(setupMultiSelect);
      document.getElementById("fromDate").addEventListener("change", e => { state.from = e.target.value; render(); });
      document.getElementById("toDate").addEventListener("change", e => { state.to = e.target.value; render(); });
      document.getElementById("cancelFilter").addEventListener("change", e => { state.cancel = e.target.value; render(); });
      document.getElementById("rawPeriodFilter").addEventListener("change", e => { state.rawPeriod = e.target.value; render(); });
      document.querySelectorAll("[data-trend-view]").forEach(btn => btn.addEventListener("click", () => { state.trendView = btn.dataset.trendView; document.querySelectorAll("[data-trend-view]").forEach(b => b.classList.toggle("active", b === btn)); render(); }));
      document.getElementById("downloadSkuData").addEventListener("click", downloadSkuData);
      document.getElementById("downloadGroupData").addEventListener("click", downloadGroupData);
      document.getElementById("downloadRawData").addEventListener("click", downloadRawData);
      document.addEventListener("click", handleOutsideClick);
      render();
    }
    function setupMultiSelect(name) {
      const root = document.getElementById(`${name}Filter`), trigger = document.getElementById(`${name}Trigger`), search = document.getElementById(`${name}Search`);
      trigger.addEventListener("click", () => { document.querySelectorAll(".multi-select.open").forEach(el => { if (el !== root) el.classList.remove("open"); }); root.classList.toggle("open"); if (root.classList.contains("open")) search.focus(); });
      search.addEventListener("input", () => renderMultiSelectOptions(name));
      root.querySelectorAll("[data-action]").forEach(btn => btn.addEventListener("click", () => { state[name] = btn.dataset.action === "all" ? [...multiSelectConfig[name].values] : []; renderMultiSelect(name); render(); }));
      renderMultiSelect(name);
    }
    function handleOutsideClick(event) { document.querySelectorAll(".multi-select.open").forEach(root => { if (!root.contains(event.target)) root.classList.remove("open"); }); }
    function renderMultiSelect(name) {
      renderMultiSelectOptions(name);
      const cfg = multiSelectConfig[name], selected = state[name], trigger = document.getElementById(`${name}Trigger`), chips = document.getElementById(`${name}Chips`);
      if (!selected.length) { trigger.textContent = cfg.allLabel; chips.innerHTML = ""; return; }
      trigger.textContent = selected.length <= 2 ? selected.join(", ") : `${selected.length} ${cfg.label} đã chọn`;
      chips.innerHTML = selected.slice(0, 6).map(value => `<span class="selected-chip">${escapeHtml(value)}<button type="button" data-value="${escapeAttr(value)}">×</button></span>`).join("");
      chips.querySelectorAll("button").forEach(btn => btn.addEventListener("click", () => { state[name] = state[name].filter(item => item !== btn.dataset.value); renderMultiSelect(name); render(); }));
    }
    function renderMultiSelectOptions(name) {
      const cfg = multiSelectConfig[name], search = document.getElementById(`${name}Search`), holder = document.getElementById(`${name}Options`), keyword = normalizeText(search.value);
      const filtered = cfg.values.filter(value => !keyword || normalizeText(value).includes(keyword)).slice(0, 240);
      holder.innerHTML = filtered.map(value => `<label class="multi-select-option"><input type="checkbox" ${state[name].includes(value) ? "checked" : ""} data-value="${escapeAttr(value)}" /><span>${escapeHtml(value)}</span></label>`).join("") || `<div class="empty">Không có dữ liệu.</div>`;
      holder.querySelectorAll("input").forEach(input => input.addEventListener("change", () => { const set = new Set(state[name]); input.checked ? set.add(input.dataset.value) : set.delete(input.dataset.value); state[name] = [...set]; renderMultiSelect(name); render(); }));
    }
    function escapeHtml(text) { return String(text ?? "").replaceAll("&","&amp;").replaceAll("<","&lt;").replaceAll(">","&gt;").replaceAll('"',"&quot;"); }
    function escapeAttr(text) { return escapeHtml(text).replaceAll("'","&#39;"); }
    function normalizeText(value) { return String(value || "").trim().toLowerCase(); }
    function imageSrc(row) { return REPORT_DATA.meta.imageAssets?.[row.sku] || REPORT_DATA.meta.imageAssets?.[row.barcode] || row.image || ""; }
    function matches(candidate, selected) { if (!selected.length) return true; const target = normalizeText(candidate); return selected.some(token => target === normalizeText(token)); }
    function inRange(date, from, to) { return date >= from && date <= to; }
    function filterRecords(source, from, to) { return source.filter(item => inRange(item.d, from, to) && matches(item.c, state.channel) && matches(item.g, state.group) && matches(item.s, state.sku) && matches(item.ks, state.keySummer) && matches(item.cl, state.classify) && (state.cancel === "all" || item.x === state.cancel)); }
    function formatNumber(num, fixed = 1) { const value = Number(num || 0); const digits = Math.abs(value) >= 100 ? 0 : fixed; return value.toLocaleString("vi-VN", { minimumFractionDigits: digits, maximumFractionDigits: digits }); }
    function formatCurrency(num) { const value = Number(num || 0); if (Math.abs(value) >= 1e9) return `${formatNumber(value / 1e9)} Tỷ`; if (Math.abs(value) >= 1e6) return `${formatNumber(value / 1e6)} Tr`; if (Math.abs(value) >= 1e3) return `${formatNumber(value / 1e3)} N`; return formatNumber(value); }
    function formatUnits(num) { const value = Number(num || 0); if (Math.abs(value) >= 1e6) return `${formatNumber(value / 1e6)} Tr`; if (Math.abs(value) >= 1e3) return `${formatNumber(value / 1e3)} N`; return formatNumber(value, 0); }
    function formatFull(num) { return Number(num || 0).toLocaleString("vi-VN", { maximumFractionDigits: 2 }); }
    function formatPercent(value) { return `${formatNumber(value, 1)}%`; }
    function pctDelta(current, previous) { if (!previous && !current) return 0; if (!previous) return 100; return ((current - previous) / previous) * 100; }
    function deltaHtml(value, inverse = false) { const cls = Math.abs(value) < .05 ? "flat" : (value > 0 ? (inverse ? "down" : "up") : (inverse ? "up" : "down")); const arrow = Math.abs(value) < .05 ? "•" : (value > 0 ? "↑" : "↓"); return `<span class="delta ${cls}">${arrow} ${formatPercent(Math.abs(value))}</span>`; }
    function addDays(iso, days) { const dt = new Date(iso + "T00:00:00"); dt.setDate(dt.getDate() + days); return dt.toISOString().slice(0, 10); }
    function dayDiff(from, to) { return Math.floor((new Date(to) - new Date(from)) / 86400000) + 1; }
    function shiftMonthClamped(isoDate, monthOffset) { const [year, month, day] = isoDate.split("-").map(Number); const targetIndex = year * 12 + month - 1 + monthOffset; const y = Math.floor(targetIndex / 12), m = ((targetIndex % 12) + 12) % 12 + 1; const last = new Date(y, m, 0).getDate(); return `${y}-${String(m).padStart(2,"0")}-${String(Math.min(day,last)).padStart(2,"0")}`; }
    function periodLabel(from, to) { return `${from.split("-").reverse().join("/")} – ${to.split("-").reverse().join("/")}`; }
    function summarize(records) { let revenue = 0, volume = 0, canceled = 0; for (const item of records) { revenue += item.r; volume += item.q; if (String(item.x).toLowerCase() === "yes") canceled += 1; } return { revenue, volume, asp: volume ? revenue / volume : 0, cancelRate: records.length ? canceled / records.length * 100 : 0 }; }
    function setMetric(valueId, deltaId, current, previous, formatter, inverse, previousFrom, previousTo) { document.getElementById(valueId).textContent = formatter(current); const delta = pctDelta(current, previous); const el = document.getElementById(deltaId); el.className = `delta ${Math.abs(delta) < .05 ? "flat" : (delta > 0 ? (inverse ? "down" : "up") : (inverse ? "up" : "down"))}`; el.innerHTML = `${Math.abs(delta) < .05 ? "•" : (delta > 0 ? "↑" : "↓")} ${formatPercent(Math.abs(delta))} vs ${periodLabel(previousFrom, previousTo)}`; }
    function aggregateBy(records, keyField) { const map = new Map(); for (const item of records) { const key = item[keyField]; const row = map.get(key) || { key, revenue: 0, volume: 0, count: 0, channelRevenue: { shopee: 0, tiktokshop: 0, web: 0 } }; row.revenue += item.r; row.volume += item.q; row.count += 1; if (row.channelRevenue[item.c] !== undefined) row.channelRevenue[item.c] += item.r; map.set(key, row); } return Array.from(map.values()); }
    function aggregateSku(records) { const map = new Map(); for (const item of records) { const key = item.s; const row = map.get(key) || { sku: key, variant: item.v, product: item.p, group: item.g, keySummer: item.ks, classify: item.cl, image: item.i, url: item.u, revenue: 0, volume: 0, count: 0, channelRevenue: { shopee: 0, tiktokshop: 0, web: 0 } }; row.revenue += item.r; row.volume += item.q; row.count += 1; if (row.channelRevenue[item.c] !== undefined) row.channelRevenue[item.c] += item.r; if (!row.product || row.product.length < item.p.length) row.product = item.p; if (!row.image && item.i) row.image = item.i; if (!row.url && item.u) row.url = item.u; map.set(key, row); } return Array.from(map.values()); }
    function sortByRevenue(items) { return items.slice().sort((a,b) => b.revenue - a.revenue); }
    function getDateParts(isoDate) { const dt = new Date(isoDate + "T00:00:00"); const year = dt.getFullYear(), month = dt.getMonth() + 1; const tmp = new Date(dt); const weekday = tmp.getDay() || 7; tmp.setDate(tmp.getDate() + 4 - weekday); const weekYear = tmp.getFullYear(); const yearStart = new Date(weekYear, 0, 1); const week = Math.ceil((((tmp - yearStart) / 86400000) + 1) / 7); return { year, month, week, weekYear }; }
    function trendPeriodInfo(iso) { const p = getDateParts(iso); if (state.trendView === "month") return { key: `${p.year}-${String(p.month).padStart(2,"0")}`, label: `${p.year}-${String(p.month).padStart(2,"0")}` }; if (state.trendView === "week") return { key: `${p.weekYear}-W${String(p.week).padStart(2,"0")}`, label: `${p.weekYear} W${p.week}` }; return { key: iso, label: iso.slice(8,10) + "/" + iso.slice(5,7) }; }
    function aggregateTrendRecords(records) { const map = new Map(); for (const item of records) { const period = trendPeriodInfo(item.d); const row = map.get(period.key) || { key: period.key, label: period.label, revenue: 0, volume: 0 }; row.revenue += item.r; row.volume += item.q; map.set(period.key, row); } return Array.from(map.values()).sort((a,b) => a.key.localeCompare(b.key)); }
    function render() { const current = filterRecords(REPORT_DATA.records, state.from, state.to); const previousFrom = shiftMonthClamped(state.from, -1), previousTo = shiftMonthClamped(state.to, -1); const previous = filterRecords(REPORT_DATA.records, previousFrom, previousTo); document.getElementById("comparisonNote").innerHTML = `<strong>Kỳ đang xem:</strong> ${periodLabel(state.from, state.to)} &nbsp;•&nbsp; <strong>So sánh cùng ngày tháng trước:</strong> ${periodLabel(previousFrom, previousTo)}`; const s = summarize(current), p = summarize(previous); setMetric("metricRevenue","metricRevenueDelta",s.revenue,p.revenue,formatCurrency,false,previousFrom,previousTo); setMetric("metricVolume","metricVolumeDelta",s.volume,p.volume,formatUnits,false,previousFrom,previousTo); setMetric("metricAsp","metricAspDelta",s.asp,p.asp,formatCurrency,false,previousFrom,previousTo); setMetric("metricCancel","metricCancelDelta",s.cancelRate,p.cancelRate,formatPercent,true,previousFrom,previousTo); renderTrendChart(current); renderBreakdown("channel","c",current,previous); renderBreakdown("group","g",current,previous); renderBreakdown("keySummer","ks",current,previous); renderBreakdown("classify","cl",current,previous); renderSkuTable(current, previous); renderRawTable(current); }
    function renderTrendChart(records) { const rows = aggregateTrendRecords(records); if (!rows.length) { document.getElementById("trendChart").innerHTML = `<div class="empty">Không có dữ liệu.</div>`; return; } const width = Math.max(960, rows.length * 34), height = 420, pad = {top:18,right:56,bottom:72,left:60}; const chartW = width - pad.left - pad.right, chartH = height - pad.top - pad.bottom; const maxRevenue = Math.max(...rows.map(r => r.revenue), 1), maxVolume = Math.max(...rows.map(r => r.volume), 1); const step = chartW / rows.length, barW = Math.max(10, step * .72); let bars = "", labels = "", path = "", dots = "", grid = ""; for (let i = 0; i <= 5; i++) { const y = pad.top + chartH * i / 5; grid += `<line x1="${pad.left}" y1="${y}" x2="${width-pad.right}" y2="${y}" stroke="#dbe6f0" />`; } rows.forEach((row, idx) => { const cx = pad.left + idx * step + step / 2, h = row.revenue / maxRevenue * chartH, x = cx - barW / 2, y = pad.top + chartH - h; bars += `<rect x="${x}" y="${y}" width="${barW}" height="${h}" rx="4" fill="#4d89e8"><title>${row.label} | DT ${formatFull(row.revenue)} | Volume ${formatFull(row.volume)}</title></rect>`; const vy = pad.top + chartH - row.volume / maxVolume * chartH; path += `${idx === 0 ? "M" : "L"} ${cx} ${vy} `; dots += `<circle cx="${cx}" cy="${vy}" r="4.5" fill="#f39a3f"><title>${row.label} | Volume ${formatFull(row.volume)}</title></circle>`; if (idx % Math.ceil(rows.length / 16) === 0 || rows.length <= 16) labels += `<text x="${cx}" y="${height-34}" text-anchor="end" transform="rotate(-35 ${cx} ${height-34})" font-size="11" fill="#60758d">${row.label}</text>`; }); document.getElementById("trendChart").innerHTML = `<svg viewBox="0 0 ${width} ${height}" role="img" aria-label="DT và volume theo kỳ">${grid}<line x1="${pad.left}" y1="${pad.top+chartH}" x2="${width-pad.right}" y2="${pad.top+chartH}" stroke="#9fb3c8" />${bars}<path d="${path}" fill="none" stroke="#f39a3f" stroke-width="3" stroke-linejoin="round" stroke-linecap="round"></path>${dots}${labels}</svg>`; }
    function renderBreakdown(prefix, keyField, current, previous) { const rows = sortByRevenue(aggregateBy(current, keyField)); const prevMap = new Map(aggregateBy(previous, keyField).map(row => [row.key, row])); document.getElementById(prefix + "Viz").innerHTML = renderSummaryViz(rows); document.getElementById(prefix + "Table").innerHTML = renderSummaryTable(rows, prevMap, prefix); attachSummarySortHandlers(prefix, rows, prevMap); }
    function renderSummaryViz(rows) { if (!rows.length) return `<div class="empty">Không có dữ liệu.</div>`; const total = rows.reduce((sum,row) => sum + row.revenue, 0); return `<div class="viz-list">${rows.slice(0,8).map((row, idx) => { const share = total ? row.revenue / total * 100 : 0; return `<div class="viz-row"><div class="viz-name">${escapeHtml(row.key)}</div><div class="viz-bar"><div class="viz-fill" style="width:${share}%; background:${PALETTE[idx % PALETTE.length]}"></div></div><div class="viz-meta">${formatCurrency(row.revenue)} · ${formatPercent(share)}</div></div>`; }).join("")}</div>`; }
    function sortSummaryRows(rows, type) { const sort = state.summarySort[type] || {key:"revenue", dir:"desc"}, factor = sort.dir === "asc" ? 1 : -1; return rows.slice().sort((a,b) => { if (sort.key === "label") return String(a.key).localeCompare(String(b.key)) * factor; const av = sort.key === "asp" ? (a.volume ? a.revenue / a.volume : 0) : (a[sort.key] || 0); const bv = sort.key === "asp" ? (b.volume ? b.revenue / b.volume : 0) : (b[sort.key] || 0); return (av - bv) * factor; }); }
    function renderSummaryTable(rows, prevMap, type) { if (!rows.length) return `<div class="empty">Không có dữ liệu.</div>`; const sorted = sortSummaryRows(rows, type), totalRevenue = rows.reduce((s,r) => s+r.revenue,0), totalVolume = rows.reduce((s,r) => s+r.volume,0); const titleMap = { channel:"Kênh bán hàng", group:"Group", keySummer:"Priority", classify:"Classify" }; const visible = type === "group" ? sorted : sorted.slice(0, 12); const sort = state.summarySort[type]; const arrow = key => sort.key === key ? (sort.dir === "desc" ? "▼" : "▲") : "↕"; const body = visible.map((row, idx) => { const prev = prevMap.get(row.key) || {revenue:0, volume:0}; const asp = row.volume ? row.revenue / row.volume : 0, prevAsp = prev.volume ? prev.revenue / prev.volume : 0, share = totalRevenue ? row.revenue / totalRevenue * 100 : 0; return `<tr><td class="rank">${idx+1}.</td><td>${escapeHtml(row.key)}<div><span class="subtle">${formatPercent(share)} DT</span></div></td><td>${formatCurrency(row.revenue)}<div>${deltaHtml(pctDelta(row.revenue, prev.revenue))}</div></td><td>${formatUnits(row.volume)}<div>${deltaHtml(pctDelta(row.volume, prev.volume))}</div></td><td>${formatCurrency(asp)}<div>${deltaHtml(pctDelta(asp, prevAsp))}</div></td></tr>`; }).join(""); return `<div class="table-scroll"><table><thead><tr><th></th><th><button class="sort-btn" data-summary-sort="${type}" data-sort-key="label">${titleMap[type]} ${arrow("label")}</button></th><th><button class="sort-btn" data-summary-sort="${type}" data-sort-key="revenue">DT ${arrow("revenue")}</button></th><th><button class="sort-btn" data-summary-sort="${type}" data-sort-key="volume">Volume ${arrow("volume")}</button></th><th><button class="sort-btn" data-summary-sort="${type}" data-sort-key="asp">ASP ${arrow("asp")}</button></th></tr></thead><tbody>${body}</tbody></table></div><div class="summary-total"><div class="summary-total-grid"><div class="summary-stat-value">Tổng cộng</div><div class="summary-stat"><div class="summary-stat-label">DT</div><div class="summary-stat-value">${formatCurrency(totalRevenue)}</div></div><div class="summary-stat"><div class="summary-stat-label">Volume</div><div class="summary-stat-value">${formatUnits(totalVolume)}</div></div><div class="summary-stat"><div class="summary-stat-label">ASP</div><div class="summary-stat-value">${formatCurrency(totalVolume ? totalRevenue / totalVolume : 0)}</div></div></div></div>`; }
    function attachSummarySortHandlers(type, rows, prevMap) { document.querySelectorAll(`[data-summary-sort="${type}"]`).forEach(btn => btn.addEventListener("click", () => { const key = btn.dataset.sortKey, cur = state.summarySort[type]; state.summarySort[type] = { key, dir: cur.key === key && cur.dir === "desc" ? "asc" : "desc" }; document.getElementById(type + "Table").innerHTML = renderSummaryTable(rows, prevMap, type); attachSummarySortHandlers(type, rows, prevMap); })); }
    function sortSkuRows(rows) { const sort = state.skuSort, factor = sort.dir === "asc" ? 1 : -1; return rows.slice().sort((a,b) => { if (["barcode","product","group"].includes(sort.key)) { const av = sort.key === "barcode" ? a.sku : a[sort.key]; const bv = sort.key === "barcode" ? b.sku : b[sort.key]; return String(av || "").localeCompare(String(bv || "")) * factor; } const av = sort.key === "asp" ? (a.volume ? a.revenue / a.volume : 0) : (a[sort.key] || 0); const bv = sort.key === "asp" ? (b.volume ? b.revenue / b.volume : 0) : (b[sort.key] || 0); return (av - bv) * factor; }); }
    function renderSkuTable(current, previous) { const rows = aggregateSku(current), prevMap = new Map(aggregateSku(previous).map(row => [row.sku, row])); if (!rows.length) { ["skuTable","skuGrowthTable","skuReduceTable"].forEach(id => document.getElementById(id).innerHTML = `<div class="empty">Không có dữ liệu.</div>`); return; } document.getElementById("skuTable").innerHTML = renderSkuTableMarkup(rows, prevMap); document.getElementById("skuGrowthTable").innerHTML = renderSkuTrendTable(rows, prevMap, "growth"); document.getElementById("skuReduceTable").innerHTML = renderSkuTrendTable(rows, prevMap, "reduce"); attachSkuSortHandlers(rows, prevMap); }
    function attachSkuSortHandlers(rows, prevMap) { document.querySelectorAll("[data-sku-sort]").forEach(btn => btn.addEventListener("click", () => { const key = btn.dataset.sortKey, cur = state.skuSort; state.skuSort = { key, dir: cur.key === key && cur.dir === "desc" ? "asc" : "desc" }; document.getElementById("skuTable").innerHTML = renderSkuTableMarkup(rows, prevMap); attachSkuSortHandlers(rows, prevMap); })); }
    function renderSkuTableMarkup(rows, prevMap) { const sorted = sortSkuRows(rows).slice(0, 100), sort = state.skuSort, arrow = key => sort.key === key ? (sort.dir === "desc" ? "▼" : "▲") : "↕"; const body = sorted.map((row, idx) => { const prev = prevMap.get(row.sku) || {revenue:0, volume:0}; const asp = row.volume ? row.revenue / row.volume : 0, prevAsp = prev.volume ? prev.revenue / prev.volume : 0; const name = row.url ? `<a class="product-link" href="${escapeAttr(row.url)}" target="_blank" rel="noopener">${escapeHtml(row.product)}</a>` : escapeHtml(row.product); const img = imageSrc(row); return `<tr><td class="rank">${idx+1}.</td><td>${img ? `<img class="sku-thumb" src="${escapeAttr(img)}" alt="${escapeAttr(row.product)}" loading="lazy" />` : ""}</td><td><span class="sku-code">${escapeHtml(row.sku)}</span><div class="subtle">${escapeHtml(row.variant || "-")}</div></td><td class="raw-product">${name}<div class="subtle">${escapeHtml(row.keySummer)} · ${escapeHtml(row.classify)}</div></td><td>${escapeHtml(row.group)}</td><td>${formatCurrency(row.revenue)}<div>${deltaHtml(pctDelta(row.revenue, prev.revenue))}</div></td><td>${formatUnits(row.volume)}<div>${deltaHtml(pctDelta(row.volume, prev.volume))}</div></td><td>${formatCurrency(asp)}<div>${deltaHtml(pctDelta(asp, prevAsp))}</div></td></tr>`; }).join(""); return `<div class="table-scroll"><table><thead><tr><th></th><th>Ảnh</th><th><button class="sort-btn" data-sku-sort data-sort-key="barcode">Barcode ${arrow("barcode")}</button></th><th><button class="sort-btn" data-sku-sort data-sort-key="product">Tên sản phẩm ${arrow("product")}</button></th><th><button class="sort-btn" data-sku-sort data-sort-key="group">Group ${arrow("group")}</button></th><th><button class="sort-btn" data-sku-sort data-sort-key="revenue">DT ${arrow("revenue")}</button></th><th><button class="sort-btn" data-sku-sort data-sort-key="volume">Volume ${arrow("volume")}</button></th><th><button class="sort-btn" data-sku-sort data-sort-key="asp">ASP ${arrow("asp")}</button></th></tr></thead><tbody>${body}</tbody></table></div>`; }
    function getSkuTrendRows(rows, prevMap, direction) { const periodDays = Math.max(dayDiff(state.from, state.to), 1); return rows.map(row => { const prev = prevMap.get(row.sku) || {revenue:0}; return {...row, deltaRevenue: pctDelta(row.revenue, prev.revenue), weeklyRevenue: row.revenue / periodDays * 7}; }).filter(row => row.weeklyRevenue > 5000000).filter(row => direction === "growth" ? row.deltaRevenue > 0 : row.deltaRevenue < 0).sort((a,b) => direction === "growth" ? b.deltaRevenue - a.deltaRevenue : a.deltaRevenue - b.deltaRevenue).slice(0,20); }
    function renderSkuTrendTable(rows, prevMap, direction) { const trendRows = getSkuTrendRows(rows, prevMap, direction); if (!trendRows.length) return `<div class="empty">Không có dữ liệu.</div>`; const body = trendRows.map((row, idx) => { const img = imageSrc(row); return `<tr><td class="rank">${idx+1}.</td><td>${img ? `<img class="sku-thumb" src="${escapeAttr(img)}" alt="${escapeAttr(row.product)}" loading="lazy" />` : ""}</td><td><span class="sku-code">${escapeHtml(row.sku)}</span><div class="subtle">${escapeHtml(row.product)}</div></td><td>${escapeHtml(row.group)}</td><td>${formatCurrency(row.revenue)}</td><td>${deltaHtml(row.deltaRevenue)}</td></tr>`; }).join(""); return `<div class="table-scroll"><table><thead><tr><th></th><th>Ảnh</th><th>SKU</th><th>Group</th><th>DT</th><th>% Δ</th></tr></thead><tbody>${body}</tbody></table></div>`; }
    function rawPeriodInfo(iso) { const p = getDateParts(iso); if (state.rawPeriod === "day") return { key: iso, label: iso, year: p.year }; if (state.rawPeriod === "month") return { key: `${p.year}-${String(p.month).padStart(2,"0")}`, label: `tháng ${p.month}`, year: p.year }; if (state.rawPeriod === "year") return { key: String(p.year), label: String(p.year), year: p.year }; return { key: `${p.weekYear}-W${String(p.week).padStart(2,"0")}`, label: String(p.week), year: p.weekYear }; }
    function aggregateRawData(records) { const map = new Map(); for (const item of records) { const period = rawPeriodInfo(item.d), key = [item.s,item.c,period.key].join("||"); const row = map.get(key) || { image:item.i, url:item.u, barcode:item.s, product:item.p, group:item.g, priority:item.ks, classify:item.cl, orderDateLabel:period.label, orderYear:period.year, channel:item.c, revenue:0, volume:0 }; row.revenue += item.r; row.volume += item.q; if (!row.image && item.i) row.image = item.i; if (!row.url && item.u) row.url = item.u; map.set(key,row); } return Array.from(map.values()).map(row => ({...row, asp: row.volume ? row.revenue / row.volume : 0})).sort((a,b) => b.revenue - a.revenue); }
    function renderRawTable(current) { const rows = aggregateRawData(current).slice(0, 600); if (!rows.length) { document.getElementById("rawTable").innerHTML = `<div class="empty">Không có dữ liệu.</div>`; return; } const periodLabel = state.rawPeriod === "month" ? "Order Date (Tháng)" : state.rawPeriod === "year" ? "Order Date (Năm)" : state.rawPeriod === "day" ? "Ngày đặt hàng" : "Order Date (Tuần)"; const body = rows.map((row, idx) => { const name = row.url ? `<a class="product-link" href="${escapeAttr(row.url)}" target="_blank" rel="noopener">${escapeHtml(row.product)}</a>` : escapeHtml(row.product); const img = imageSrc(row); return `<tr><td class="rank">${idx+1}.</td><td>${img ? `<img class="raw-thumb" src="${escapeAttr(img)}" alt="${escapeAttr(row.product)}" loading="lazy" />` : ""}</td><td><span class="sku-code">${escapeHtml(row.barcode)}</span></td><td class="raw-product">${name}</td><td>${escapeHtml(row.group)}</td><td>${escapeHtml(row.priority)}</td><td>${escapeHtml(row.classify)}</td><td>${escapeHtml(row.orderDateLabel)}</td><td>${escapeHtml(String(row.orderYear))}</td><td>${escapeHtml(row.channel)}</td><td>${formatCurrency(row.revenue)}</td><td>${formatUnits(row.volume)}</td><td>${formatCurrency(row.asp)}</td></tr>`; }).join(""); document.getElementById("rawTable").innerHTML = `<div class="table-scroll raw-table"><table><thead><tr><th></th><th>Ảnh</th><th>Barcode</th><th>Tên sản phẩm</th><th>Group</th><th>Priority</th><th>Classify</th><th>${periodLabel}</th><th>Năm</th><th>Kênh bán hàng</th><th>DT</th><th>Volume</th><th>ASP</th></tr></thead><tbody>${body}</tbody></table></div>`; }
    function buildCsv(header, rows) { return "\\uFEFF" + [header].concat(rows).map(cols => cols.map(value => `"${String(value ?? "").replaceAll('"','""')}"`).join(",")).join("\\r\\n"); }
    function downloadFile(name, csv) { const blob = new Blob([csv], {type:"text/csv;charset=utf-8;"}); const url = URL.createObjectURL(blob); const a = document.createElement("a"); a.href = url; a.download = name; a.style.display = "none"; document.body.appendChild(a); a.click(); a.remove(); URL.revokeObjectURL(url); }
    function downloadSkuData() { const rows = sortByRevenue(aggregateSku(filterRecords(REPORT_DATA.records, state.from, state.to))); downloadFile("sku_detail_export.csv", buildCsv(["image","product_url","barcode","variant_id","product_name","group","priority","classify","revenue","volume","asp"], rows.map(row => [row.image || "",row.url || "",row.sku,row.variant,row.product,row.group,row.keySummer,row.classify,row.revenue,row.volume,row.volume ? row.revenue / row.volume : 0]))); }
    function downloadGroupData() { const rows = sortByRevenue(aggregateBy(filterRecords(REPORT_DATA.records, state.from, state.to), "g")); const total = rows.reduce((s,r) => s + r.revenue, 0); downloadFile("group_performance_export.csv", buildCsv(["group","revenue","volume","asp","share"], rows.map(row => [row.key,row.revenue,row.volume,row.volume ? row.revenue / row.volume : 0,total ? row.revenue / total * 100 : 0]))); }
    function downloadRawData() { const rows = aggregateRawData(filterRecords(REPORT_DATA.records, state.from, state.to)); downloadFile(`order_haravan_raw_${state.rawPeriod}.csv`, buildCsv(["image","product_url","barcode","product_name","group","priority","product_classify","order_period","order_year","sales_channel","revenue","volume","asp"], rows.map(row => [row.image || "",row.url || "",row.barcode,row.product,row.group,row.priority,row.classify,row.orderDateLabel,row.orderYear,row.channel,row.revenue,row.volume,row.asp]))); }
    setup();
  </script>
</body>
</html>""".replace("__REPORT_JSON__", report_json)


def main():
    records = read_records()
    report_json = json.dumps(build_report_data(records), ensure_ascii=False, separators=(",", ":"))
    OUT.write_text(html_template(report_json), encoding="utf-8")
    print(f"Wrote {OUT}")
    print(f"Records: {len(records):,}")


if __name__ == "__main__":
    main()
